"""Convert data/_full_workbook.xlsx -> data/<analyst>/messages.{jsonl,csv}

One JSONL and one CSV per trader tab, sorted by timestamp ascending. Slim schema:
    analyst, timestamp, content
"""
import csv
import json
from pathlib import Path

import openpyxl

FIELDS = ["analyst", "timestamp", "content"]

DATA = Path(__file__).resolve().parent.parent / "data"
SRC = DATA / "_full_workbook.xlsx"

# Tabs we treat as trader corpora. Other tabs (Enhanced Market, Sheet8, *Gaps)
# have different schemas and need a separate decision before ingest.
TRADER_TABS = ["ECS", "Eva", "Grizzlies", "Waxui", "Zabes", "Nando", "Ace"]

KEEP = ["content", "is_reply", "reply_to_message_id"]


def normalize(value):
    if value is None:
        return ""
    return str(value)


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    summary = []
    for tab in TRADER_TABS:
        if tab not in wb.sheetnames:
            print(f"  SKIP {tab!r}: tab not present in workbook")
            continue
        ws = wb[tab]
        rows = ws.iter_rows(values_only=True)
        header = [h for h in next(rows)]
        idx = {name: i for i, name in enumerate(header)}

        records = []
        for row in rows:
            if row is None or all(c is None for c in row):
                continue
            timestamp = normalize(row[idx["timestamp"]]) if "timestamp" in idx else ""
            content = normalize(row[idx["content"]]) if "content" in idx else ""
            if not timestamp and not content:
                continue
            records.append({
                "analyst": tab,
                "timestamp": timestamp,
                "content": content,
            })

        out_dir = DATA / tab
        out_dir.mkdir(parents=True, exist_ok=True)
        jsonl_file = out_dir / "messages.jsonl"

        # Merge with existing messages (preserves Discord exports)
        existing = {}
        if jsonl_file.exists():
            for line in jsonl_file.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line:
                    m = json.loads(line)
                    existing[(m["timestamp"], m["content"])] = m

        for r in records:
            existing[(r["timestamp"], r["content"])] = r

        merged = sorted(existing.values(), key=lambda r: r["timestamp"])

        with jsonl_file.open("w", encoding="utf-8") as f:
            for r in merged:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")

        csv_file = out_dir / "messages.csv"
        with csv_file.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDS)
            writer.writeheader()
            writer.writerows(merged)

        records = merged

        summary.append((tab, len(records), jsonl_file))
        print(f"  {tab:15s} -> data/{tab}/messages.{{jsonl,csv}}  ({len(records)} messages)")

    print()
    print(f"total messages: {sum(n for _, n, _ in summary)} across {len(summary)} traders")


if __name__ == "__main__":
    main()
